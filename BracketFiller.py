# -*- coding: utf-8 -*-
"""
Created on Sat Apr 18 17:29:48 2015

@author: Ron
"""

import Competitor
import copy
import math
from openpyxl import Workbook, load_workbook

# Assume that bracket numbers are ordered as a binary tree
#       1
#     2  3
#   4 5 6 7
# ...

# Maps the fighter number to their position in the bracket
fighter_to_excel_map = [['G16', # Winner of 2 person bracket
                         'D11', 'D21'],
                        ['J16', # Winner of 4 person bracket
                         'H10', 'H22', # Finals
                         'E7', 'E13', 'E19', 'E25'],
                        ['L18', # Winner of 8 person bracket
                         'J10', 'J26', #Finals
                         'I6', 'I13', 'I23', 'I30', #Semifinals
                         'F4', 'F8', 'F11', 'F15', 'F21', 'F25', 'F28', 'F32']]                                   
# Maps some written information to excel cell
info_to_excel_map = {'gender': 'G1', 'age': 'H1', 'belt': 'I1', 'weight': 'J1'}

# Age groups
age_groups = [7, 11, 14, 17, 32]
# Weight groups
men_weight = [135.5, 150.5, 165.5, 180.5, 195.5]
women_weight = [108.5, 120.5, 135.5, 150.5, 165.5]

bye = Competitor.ByeCompetitor()

### GENERAL UTILITIES
def read_str_val(ws, cell):
    val = ws[cell].value
    if (val != None):
        return val
    else:
        return ''

def read_num_val(ws, cell):
    val = ws[cell].value
    try:
        return float(val)
    except:
        return 0

def get_gender_from_str(text):
    gender = text.strip().lower()[0]
    if (gender != 'm' and gender != 'f'):
        return ''
    return gender

def get_belt_from_str(text):
    belt = text.strip().lower()
    if (belt in Competitor.SparringCompetitor.Belts):
        return belt
    else:
        return None

### BRACKET RELATED CODE
def get_bracket_gender(competitors):
    for c in competitors:
        if (c.gender != ''):
            return c.gender
    return ''

def get_bracket_age(competitors):
    s = 0
    for c in competitors:
        if (c.age > 0):
            s += c.age
    age = s/len(competitors)
    # Do linear search since there are so few age groups
    for i in range(len(age_groups)):
        if (age < age_groups[i]):
            break
    
    if (i == 0):
        return '6-7'
    if (i == len(competitors)):
        return '33+'
    return str(age_groups[i-1]) + '-' + str(age_groups[i])

def get_bracket_belt(competitors):
    for c in competitors:
        if (c.belt != None):
            belt = c.belt

    if (belt == Competitor.SparringCompetitor.Belts[0] \
        or belt == Competitor.SparringCompetitor.Belts[1]): # Yellow or Green
        return 'Yellow/Green'
    if (belt == Competitor.SparringCompetitor.Belts[2] \
        or belt == Competitor.SparringCompetitor.Belts[3]): # Blue or Red
        return 'Blue/Red'
    if (belt == Competitor.SparringCompetitor.Belts[4]): # Black
        return 'Black'
    
    return ''
        
def get_bracket_weight(competitors):
    min_weight = float("inf")
    max_weight = float("-inf")
    
    for c in competitors:
        if (c.weight <= 0):
            continue
        if (c.weight < min_weight):
            min_weight = c.weight
        if (c.weight > max_weight):
            max_weight = c.weight
    
    return str(min_weight) + '-' + str(max_weight)

def get_bracket_indices(n, node, indices, bracket_size):
    if (n == 1):
        if (node >= bracket_size):
            indices.append(int(node))
        else:
            indices.append(int(2*node))
    else:
        get_bracket_indices(math.ceil(n/2.), 2*node, indices, bracket_size)
        get_bracket_indices(math.floor(n/2.), 2*node+1, indices, bracket_size)

def get_competitors_per_division(ws, start):
    '''
        Reads worksheet ws to obtain the number of competitors in
        the division starting from row 'start'. Divisions are
        separated by empty rows.
    '''
    count = 0
    while(ws['A'+str(start+count)].value != None):
        count += 1
    return count

def construct_competitor(ws, ix):
    '''
        Constructs SparringCompetitor object from ix row of worksheet ws.
        
    '''
    c = Competitor.SparringCompetitor()
    c.first_name = read_str_val(ws, 'B' + str(ix))
    c.last_name = read_str_val(ws, 'A' + str(ix))
    c.school = read_str_val(ws, 'P' + str(ix))
    c.age = read_num_val(ws, 'H' + str(ix))
    c.gender = get_gender_from_str(read_str_val(ws, 'G' + str(ix)))
    c.weight = read_num_val(ws, 'L' + str(ix))
    c.belt = get_belt_from_str(read_str_val(ws, 'N' + str(ix)))
   
    return c

def get_competitors(sparws, start, count):
    '''
        Reads rows 'start' to 'start+count' to obtain the names
        and schools of competitors in the division defined by those
        rows. Result is stored in the list 'competitors'
    '''
    competitors = []
    for i in range(count):
        competitors.append(construct_competitor(sparws, start+i))
    return competitors

def get_cell_below(cell):
    for i in range(len(cell)-1, -1, -1):
        if (cell[:i].isalpha()):
            return cell[:i] + str(int(cell[i:])+1)

def write_competitors_to_bracket(ws, competitors, indices, outwb, output_fname, bracket_size):
    num_competitors = len(competitors)
    ctr = 0
    for i in range(bracket_size, 2*bracket_size):
        if (ctr < num_competitors and i == indices[ctr]):
            c = competitors[ctr]
            ctr += 1
        else:
            c = bye
        name = c.first_name + ' ' + c.last_name
        bracket_index = int(math.ceil(math.log(num_competitors,2)))-1;
        fighter_index = i-1
        cell = fighter_to_excel_map[bracket_index][fighter_index]
        ws[cell] = name
        ws[get_cell_below(cell)] = c.school
    outwb.save(output_fname)

def fill_first_line(ws, competitors, outwb, output_fname):
    # Write gender
    ws[info_to_excel_map['gender']] = get_bracket_gender(competitors)
    ws[info_to_excel_map['age']] = get_bracket_age(competitors)
    ws[info_to_excel_map['belt']] = get_bracket_belt(competitors)
    ws[info_to_excel_map['weight']] = get_bracket_weight(competitors)
    outwb.save(output_fname)
    

def get_bracket_template_sheet(outwb, templatewb, division, n):
    '''
        n >= 2
    '''
    bracket_size = int(2**math.ceil(math.log(n,2)))
    sheet_name = str(bracket_size) + '-Person'
    tws = templatewb.get_sheet_by_name(sheet_name)
    outwb.add_sheet(copy.deepcopy(tws), division)
    ws = outwb.worksheets[division]
    return ws

def fill_in_brackets(template_fname, data_fname, output_fname):

    templatewb = load_workbook(template_fname)

    outwb = Workbook() # Output file
    outwb.save(output_fname)

    sparwb = load_workbook(data_fname) # Bracket data
    sparws = sparwb.active
    
    row = 2;
    division = 0;
    
    while (True):
        comp_in_div = get_competitors_per_division(sparws, row)
       
        if (comp_in_div <= 0):
            break
        
        if (comp_in_div == 1):
            row += comp_in_div + 1
            continue
        
        competitors = get_competitors(sparws, row, comp_in_div)
        
        bracket_indices = []
        bracket_size = int(2**math.ceil(math.log(comp_in_div, 2)));
        get_bracket_indices(comp_in_div, 1, bracket_indices, bracket_size)
        bracket_indices.sort()
        
        outws = get_bracket_template_sheet(outwb, templatewb, division, comp_in_div)        
        outws.title = 'Division ' + str(division)        

        write_competitors_to_bracket(outws, competitors, bracket_indices, outwb, output_fname, bracket_size)
        fill_first_line(outws, competitors, outwb, output_fname)

        row += comp_in_div + 1
        division += 1
        
        print division

    outwb.save(output_fname)